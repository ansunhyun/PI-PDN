import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("❌ 오류: openpyxl 또는 Pillow 라이브러리가 설치되지 않았습니다.")
    print("CMD에서 다음 명령어를 실행하세요: pip install openpyxl Pillow")
    sys.exit()

# 1. 설정 (명령어 인수로 파일 경로 받기)
if len(sys.argv) > 1:
    JSON_FILE = sys.argv[1]
else:
    JSON_FILE = 'result.json'

# JSON 파일이 있는 폴더를 OUTPUT_DIR로 자동 설정 (예: .../outputs)
OUTPUT_DIR = os.path.dirname(JSON_FILE) 
if not OUTPUT_DIR:
    OUTPUT_DIR = 'output'

# ---------------------------------------------------------
# 💡 수정된 부분: 대소문자를 정확히 맞춘 modelInfo 구조 반영
# ---------------------------------------------------------
MODEL_NAME = 'UNKNOWN_MODEL'
PARENT_DIR = os.path.dirname(OUTPUT_DIR) # 상위 폴더

# request.json이 있을 만한 경로 후보 (outputs 폴더 내부 또는 상위 폴더)
request_json_paths = [
    os.path.join(OUTPUT_DIR, 'request.json'),
    os.path.join(PARENT_DIR, 'request.json')
]

req_data_found = False

for req_path in request_json_paths:
    if os.path.exists(req_path):
        try:
            with open(req_path, 'r', encoding='utf-8') as req_f:
                req_data = json.load(req_f)
                
                # 1순위: modelInfo -> name 구조 확인 (대문자 'I' 주의)
                if 'modelInfo' in req_data and 'name' in req_data['modelInfo']:
                    MODEL_NAME = req_data['modelInfo']['name']
                    req_data_found = True
                    break
                # 혹시 모를 소문자 modelinfo 방어 코드
                elif 'modelinfo' in req_data and 'name' in req_data['modelinfo']:
                    MODEL_NAME = req_data['modelinfo']['name']
                    req_data_found = True
                    break
                # 2순위: 최상단에 name이 있는 경우
                elif 'name' in req_data:
                    MODEL_NAME = req_data['name']
                    req_data_found = True
                    break
        except json.JSONDecodeError:
            print(f"⚠️ 경고: '{req_path}' 파일의 JSON 형식이 올바르지 않습니다.")

# 3순위 (Fallback): request.json에서 모델명을 못 찾은 경우, 상위 폴더명을 모델명으로 사용
if not req_data_found or MODEL_NAME == 'UNKNOWN_MODEL':
    fallback_name = os.path.basename(PARENT_DIR)
    if fallback_name and fallback_name.strip() != "":
        MODEL_NAME = fallback_name
        print(f"⚠️ 알림: request.json에서 모델명을 찾지 못해 폴더명 '{MODEL_NAME}'을(를) 모델명으로 사용합니다.")
    else:
        print(f"⚠️ 경고: 모델명을 찾을 수 없어 기본값({MODEL_NAME})을 사용합니다.")
# ---------------------------------------------------------

# 출력할 엑셀 파일 경로 (outputs 폴더 / 모델명_report.xlsx)
EXCEL_FILE = os.path.join(OUTPUT_DIR, f'{MODEL_NAME}_report.xlsx')

# 2. JSON 데이터 읽기 (result.json)
try:
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
except FileNotFoundError:
    print(f"❌ 오류: '{JSON_FILE}' 파일을 찾을 수 없습니다.")
    sys.exit()

summary_data = data.get("summary", [])

if not summary_data:
    print("⚠️ 경고: JSON 파일 내에 'summary' 데이터가 없습니다.")
    sys.exit()

# 3. 엑셀 워크북 생성 및 설정
wb = Workbook()
ws = wb.active
ws.title = "PDN Report"

# 헤더(Header) 작성
headers = [
    "IC", "Net", "Vmag", "Imag", "MinSpec", "MaxSpec", 
    "Result", "Drop Voltage", "Drop Rate", "Pass/Fail",
    "FitView Image", "ZoomView Image"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 이미지가 들어갈 열(K, L)의 너비 설정
ws.column_dimensions['K'].width = 45
ws.column_dimensions['L'].width = 45

# 4. 데이터 및 이미지 삽입
for row_idx, item in enumerate(summary_data, start=2):
    # 텍스트 데이터 입력
    ws.cell(row=row_idx, column=1, value=item.get('IC', '-'))
    ws.cell(row=row_idx, column=2, value=item.get('Net', '-'))
    ws.cell(row=row_idx, column=3, value=item.get('Vmag', '-'))
    ws.cell(row=row_idx, column=4, value=item.get('Imag', '-'))
    ws.cell(row=row_idx, column=5, value=item.get('MinSpec', '-'))
    ws.cell(row=row_idx, column=6, value=item.get('MaxSpec', '-'))
    ws.cell(row=row_idx, column=7, value=item.get('Result', '-'))
    ws.cell(row=row_idx, column=8, value=item.get('Drop Voltage', '-'))
    ws.cell(row=row_idx, column=9, value=item.get('Drop Rate', '-'))
    ws.cell(row=row_idx, column=10, value=item.get('Pass/Fail', '-'))
    
    # 텍스트 가운데 정렬
    for col_num in range(1, 11):
        ws.cell(row=row_idx, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    # 이미지가 들어갈 행의 높이 설정
    ws.row_dimensions[row_idx].height = 160

    # 이미지 파일명 가져오기
    fitview_filename = item.get('FitView', '')
    zoomview_filename = item.get('ZoomView', '')

    # FitView 이미지 삽입
    if fitview_filename:
        fitview_path = os.path.join(OUTPUT_DIR, fitview_filename)
        if os.path.exists(fitview_path):
            img = Image(fitview_path)
            img.width = 300  # 엑셀 셀 크기에 맞게 이미지 너비 조정
            img.height = 200 # 엑셀 셀 크기에 맞게 이미지 높이 조정
            ws.add_image(img, f"K{row_idx}")
        else:
            ws.cell(row=row_idx, column=11, value="Image Not Found").alignment = Alignment(horizontal="center", vertical="center")

    # ZoomView 이미지 삽입
    if zoomview_filename:
        zoomview_path = os.path.join(OUTPUT_DIR, zoomview_filename)
        if os.path.exists(zoomview_path):
            img = Image(zoomview_path)
            img.width = 300
            img.height = 200
            ws.add_image(img, f"L{row_idx}")
        else:
            ws.cell(row=row_idx, column=12, value="Image Not Found").alignment = Alignment(horizontal="center", vertical="center")

# 5. 엑셀 파일 저장
wb.save(EXCEL_FILE)
print(f"✅ 엑셀 리포트 생성이 완료되었습니다!")
print(f"📂 파일 위치: {EXCEL_FILE}")
