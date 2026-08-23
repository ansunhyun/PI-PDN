from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from .contracts import (
    SUPPORTED_BOM_EXPANDED_DESIGNATOR_POLICIES,
    ReferencePreprocessError,
)


DEFAULT_DESIGNATOR_COLUMNS = (
    "Designator",
    "Designators",
    "RefDes",
    "Reference",
    "Reference Designator",
    "PartReference",
    "Ref",
)

DEFAULT_PART_NUMBER_COLUMNS = (
    "PartNumber",
    "Part_Number",
    "PN",
    "Part Number",
    "Part No.",
    "Part No",
)

DEFAULT_PART_NAME_COLUMNS = (
    "PartName",
    "Part_Name",
    "Part Name",
    "Name",
)

DEFAULT_DESCRIPTION_COLUMNS = (
    "Description",
    "Desc",
    "PartDescription",
    "Part Description",
    "Site Specification",
    "SiteSpec",
    "Specification",
    "Spec",
)

DEFAULT_SYMBOL_COLUMNS = (
    "Symbol",
    "SymbolName",
    "Symbol Name",
    "Schematic Symbol",
)


def _normalized(value: object) -> str:
    return "".join(character for character in str(value).casefold() if character.isalnum())


def _rows_from_csv(path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp949"):
        try:
            with path.open("r", newline="", encoding=encoding) as stream:
                return [list(row) for row in csv.reader(stream) if any(cell.strip() for cell in row)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ReferencePreprocessError(f"unable to decode BOM CSV {path}: {last_error}")


def _rows_from_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ReferencePreprocessError(
            "XLSX BOM input requires openpyxl in the SI-TDR runtime"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
            if any(value is not None and str(value).strip() for value in row)
        ]
    finally:
        workbook.close()


def _split_designators(values: Iterable[object]) -> set[str]:
    designators: set[str] = set()
    for raw in values:
        for token in str(raw).replace(";", ",").split(","):
            value = token.strip().strip("'").strip('"')
            if value:
                designators.add(value)
    return designators


def parse_bom_rows(
    path: Path,
    *,
    configured_designator_columns: Iterable[str] = (),
) -> list[dict[str, str]]:
    """Return BOM rows keyed by their original, trimmed column names.

    The header row is located using the configured or maintained Designator
    aliases. Column names are compared case/punctuation-insensitively by
    higher-level rule loaders, but the original names remain in each row so
    validation and explain output can show the customer's BOM vocabulary.
    """

    path = path.resolve()
    configured_columns = tuple(
        str(name).strip()
        for name in configured_designator_columns
        if str(name).strip()
    )
    if not path.is_file():
        raise ReferencePreprocessError(f"BOM file not found: {path}")
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        rows = _rows_from_csv(path)
    elif suffix == ".xlsx":
        rows = _rows_from_xlsx(path)
    elif suffix == ".xls":
        raise ReferencePreprocessError(
            "legacy .xls BOM is unresolved; convert it to CSV or XLSX explicitly"
        )
    else:
        raise ReferencePreprocessError(
            f"unsupported BOM format {path.suffix!r}; expected CSV or XLSX"
        )
    if not rows:
        raise ReferencePreprocessError(f"BOM contains no rows: {path}")

    cleaned = [
        row[1:] if row and row[0].strip().endswith(":") else row
        for row in rows
    ]
    candidate_columns = configured_columns or DEFAULT_DESIGNATOR_COLUMNS
    accepted = {_normalized(name) for name in candidate_columns}
    header_index = next(
        (
            index
            for index, row in enumerate(cleaned)
            if any(_normalized(cell) in accepted for cell in row)
        ),
        None,
    )
    if header_index is None:
        raise ReferencePreprocessError(f"BOM designator column not found: {path}")

    header = [str(name).strip() for name in cleaned[header_index]]
    normalized_seen: dict[str, str] = {}
    for name in header:
        normalized = _normalized(name)
        if not normalized:
            continue
        if normalized in normalized_seen:
            raise ReferencePreprocessError(
                "BOM has duplicate normalized columns "
                f"{normalized_seen[normalized]!r} and {name!r}: {path}"
            )
        normalized_seen[normalized] = name

    parsed: list[dict[str, str]] = []
    for raw_row in cleaned[header_index + 1 :]:
        row = {
            name: (str(raw_row[index]).strip() if index < len(raw_row) else "")
            for index, name in enumerate(header)
            if name
        }
        if any(value for value in row.values()):
            parsed.append(row)
    if not parsed:
        raise ReferencePreprocessError(f"BOM contains no data rows: {path}")
    return parsed


def match_installed_bom_designator(
    component_name: str,
    installed_designators: set[str],
    expanded_policy: str,
) -> str | None:
    """Return the BOM designator that explicitly covers an imported component.

    ANF import can expand a package designator such as ``AR702`` into SIWave
    component instances ``AR702_0`` through ``AR702_3``. The expansion is
    accepted only when the Config opts into the numeric-suffix policy.
    """

    if (
        not isinstance(expanded_policy, str)
        or expanded_policy not in SUPPORTED_BOM_EXPANDED_DESIGNATOR_POLICIES
    ):
        raise ReferencePreprocessError(
            f"unsupported BOM expanded designator policy: {expanded_policy}"
        )
    if component_name in installed_designators:
        return component_name
    if expanded_policy == "parent_numeric_suffix":
        parent, separator, suffix = component_name.rpartition("_")
        if separator and parent in installed_designators and suffix.isdecimal():
            return parent
    return None


def parse_bom_designators(
    path: Path,
    *,
    configured_columns: Iterable[str] = (),
) -> set[str]:
    """Read the installed component designators without importing DCIR code.

    The parser accepts the row-oriented CSV/Part List format used by DCIR and
    `.xlsx`.  A leading `1:`/`2:` line-number cell is removed using the same
    rule as the maintained DCIR parser.  Legacy binary `.xls` is deliberately
    rejected instead of silently guessing a conversion policy.
    """

    configured_columns = tuple(
        str(name).strip() for name in configured_columns if str(name).strip()
    )
    rows = parse_bom_rows(
        path,
        configured_designator_columns=configured_columns,
    )
    candidate_columns = configured_columns or DEFAULT_DESIGNATOR_COLUMNS
    accepted = {_normalized(name) for name in candidate_columns}
    header = list(rows[0])
    normalized_header = [_normalized(name) for name in header]
    if configured_columns:
        column_indexes = [
            index for index, name in enumerate(normalized_header) if name in accepted
        ]
    else:
        column_indexes = []
        for candidate in DEFAULT_DESIGNATOR_COLUMNS:
            normalized_candidate = _normalized(candidate)
            if normalized_candidate in normalized_header:
                column_indexes = [normalized_header.index(normalized_candidate)]
                break
    selected_columns = [header[index] for index in column_indexes]
    values = (row.get(column, "") for row in rows for column in selected_columns)
    designators = _split_designators(values)
    if not designators:
        raise ReferencePreprocessError(f"BOM has no designators: {path}")
    return designators
